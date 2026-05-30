"""导出测量结果到 Excel (.xlsx)。

输出三个工作表：
- 「失效点」：编号、原始像素坐标、投影坐标、所在管线、偏离、状态
- 「距离表」：所有星-星对（含跨管线）、沿管/直线像素与米值、是否需复核
- 「标定」：全局 m/px、每条管线的 m/px

依赖：openpyxl（requirements 已含）。
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.models import FailurePoint, MeasureResult, Pipeline


HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="E0E0E0")
REVIEW_FILL = PatternFill("solid", fgColor="FFF3D6")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def export_to_xlsx(
    out_path: str | Path,
    failure_points: list[FailurePoint],
    measures: list[MeasureResult],
    pipelines: list[Pipeline],
    global_scale_m_per_px: float | None,
    per_pipeline_scale: list[float | None],
    project_name: str = "",
    drawing_name: str = "",
) -> Path:
    """写出 xlsx 文件，返回最终绝对路径。"""
    wb = Workbook()
    _write_failure_points_sheet(wb.active, failure_points, pipelines)
    _write_measures_sheet(wb.create_sheet("距离表"), measures)
    _write_calibration_sheet(
        wb.create_sheet("标定"),
        global_scale_m_per_px,
        per_pipeline_scale,
        pipelines,
        project_name,
        drawing_name,
    )
    out = Path(out_path)
    wb.save(out)
    return out.resolve()


# ---- 失效点 ----
def _write_failure_points_sheet(ws, points: list[FailurePoint], pipelines: list[Pipeline]) -> None:
    ws.title = "失效点"
    pipe_name_by_id = {p.id: p.name for p in pipelines}
    headers = ["编号", "原始 X (px)", "原始 Y (px)",
               "投影 X (px)", "投影 Y (px)",
               "管线", "偏离 (px)", "状态"]
    _write_header(ws, headers)
    for r, fp in enumerate(points, start=2):
        rx, ry = fp.raw_px
        px, py = fp.projected_px if fp.projected_px else (None, None)
        pipe = pipe_name_by_id.get(fp.pipeline_id or "", "—") if fp.pipeline_id else "—"
        row = [
            fp.code,
            round(rx, 2), round(ry, 2),
            round(px, 2) if px is not None else "—",
            round(py, 2) if py is not None else "—",
            pipe,
            round(fp.offset_px, 2) if fp.offset_px is not None else "—",
            fp.state.value,
        ]
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = CENTER
        if fp.state.value != "auto":
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = REVIEW_FILL
    _autosize(ws, headers)


# ---- 距离表 ----
def _write_measures_sheet(ws, measures: list[MeasureResult]) -> None:
    headers = ["起点", "终点", "管线",
               "沿管距离 (m)", "直线距离 (m)",
               "沿管 (px)", "直线 (px)",
               "需复核"]
    _write_header(ws, headers)
    for r, m in enumerate(measures, start=2):
        row = [
            m.from_code, m.to_code, m.pipeline_name,
            _round(m.along_pipe_m, 3) if m.along_pipe_m is not None else "—",
            _round(m.straight_m, 3) if m.straight_m is not None else "—",
            _round(m.along_pipe_px, 2),
            _round(m.straight_px, 2),
            "是" if m.need_review else "",
        ]
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = RIGHT if c >= 4 else CENTER
        if m.need_review:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = REVIEW_FILL
    _autosize(ws, headers)


# ---- 标定 ----
def _write_calibration_sheet(
    ws,
    global_scale: float | None,
    per_pipeline_scale: list[float | None],
    pipelines: list[Pipeline],
    project_name: str,
    drawing_name: str,
) -> None:
    ws.cell(row=1, column=1, value="项目").font = HEADER_FONT
    ws.cell(row=1, column=2, value=project_name or "—")
    ws.cell(row=2, column=1, value="图纸").font = HEADER_FONT
    ws.cell(row=2, column=2, value=drawing_name or "—")
    ws.cell(row=3, column=1, value="导出时间").font = HEADER_FONT
    ws.cell(row=3, column=2, value=datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"))
    ws.cell(row=4, column=1, value="全局 m/px").font = HEADER_FONT
    ws.cell(row=4, column=2, value=_round(global_scale, 6) if global_scale else "—")

    ws.cell(row=6, column=1, value="管线").font = HEADER_FONT
    ws.cell(row=6, column=2, value="m/px").font = HEADER_FONT
    ws.cell(row=6, column=1).fill = HEADER_FILL
    ws.cell(row=6, column=2).fill = HEADER_FILL
    for i, pl in enumerate(pipelines):
        ws.cell(row=7 + i, column=1, value=pl.name)
        s = per_pipeline_scale[i] if i < len(per_pipeline_scale) else None
        ws.cell(row=7 + i, column=2, value=_round(s, 6) if s else "—")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30


# ---- helpers ----
def _write_header(ws, headers: list[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _autosize(ws, headers: list[str]) -> None:
    for c in range(1, len(headers) + 1):
        col = get_column_letter(c)
        max_len = max(
            [len(str(headers[c - 1]))]
            + [len(str(ws.cell(row=r, column=c).value or "")) for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[col].width = min(28, max_len + 4)


def _round(v: float | None, digits: int) -> float | None:
    return round(v, digits) if v is not None else None
