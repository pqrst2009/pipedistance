"""导出 Excel 的最小验证测试。"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.domain.models import (
    FailurePoint,
    MeasureResult,
    Pipeline,
    ReviewState,
    new_id,
)
from app.services.export_service import export_to_xlsx


@pytest.fixture
def fixtures():
    pipelines = [
        Pipeline.from_polyline("d1", "P1", [(0.0, 0.0), (100.0, 0.0)]),
        Pipeline.from_polyline("d1", "P2", [(0.0, 200.0), (100.0, 200.0)]),
    ]
    fps = [
        FailurePoint(id=new_id("fp_"), drawing_id="d1", code="F1",
                     raw_px=(10.0, 5.0), projected_px=(10.0, 0.0),
                     pipeline_id=pipelines[0].id, offset_px=5.0,
                     marker_type="red_star", state=ReviewState.AUTO),
        FailurePoint(id=new_id("fp_"), drawing_id="d1", code="F2",
                     raw_px=(50.0, 205.0), projected_px=(50.0, 200.0),
                     pipeline_id=pipelines[1].id, offset_px=5.0,
                     marker_type="red_star", state=ReviewState.AUTO),
    ]
    measures = [
        MeasureResult(id=new_id("mr_"), drawing_id="d1",
                      from_code="F1", to_code="F2",
                      pipeline_name="—（跨管线）",
                      straight_px=204.0, along_pipe_px=0.0,
                      straight_m=20.4, along_pipe_m=None,
                      basis="cross_pipeline", need_review=True),
    ]
    return pipelines, fps, measures


def test_export_writes_three_sheets(tmp_path: Path, fixtures):
    pipelines, fps, measures = fixtures
    out = tmp_path / "result.xlsx"
    export_to_xlsx(
        out, fps, measures, pipelines,
        global_scale_m_per_px=0.1,
        per_pipeline_scale=[0.1, 0.1],
        project_name="test",
        drawing_name="d1",
    )
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["失效点", "距离表", "标定"]


def test_export_failure_points_content(tmp_path: Path, fixtures):
    pipelines, fps, measures = fixtures
    out = tmp_path / "r.xlsx"
    export_to_xlsx(out, fps, measures, pipelines, 0.1, [0.1, 0.1])
    wb = load_workbook(out)
    ws = wb["失效点"]
    # 表头 + 2 行
    assert ws.max_row == 3
    assert ws.cell(2, 1).value == "F1"
    assert ws.cell(3, 1).value == "F2"
    # 管线列要正确映射
    assert ws.cell(2, 6).value == "P1"
    assert ws.cell(3, 6).value == "P2"


def test_export_distances_content(tmp_path: Path, fixtures):
    pipelines, fps, measures = fixtures
    out = tmp_path / "r.xlsx"
    export_to_xlsx(out, fps, measures, pipelines, 0.1, [0.1, 0.1])
    wb = load_workbook(out)
    ws = wb["距离表"]
    assert ws.max_row == 2
    assert (ws.cell(2, 1).value, ws.cell(2, 2).value) == ("F1", "F2")
    assert ws.cell(2, 4).value == "—"         # 沿管米值为 None → 占位
    assert ws.cell(2, 5).value == 20.4        # 直线米值
    assert ws.cell(2, 8).value == "是"         # need_review


def test_export_calibration_content(tmp_path: Path, fixtures):
    pipelines, fps, measures = fixtures
    out = tmp_path / "r.xlsx"
    export_to_xlsx(out, fps, measures, pipelines, 0.1, [0.1, 0.2],
                   project_name="proj", drawing_name="dwg1")
    wb = load_workbook(out)
    ws = wb["标定"]
    assert ws.cell(1, 2).value == "proj"
    assert ws.cell(2, 2).value == "dwg1"
    assert ws.cell(4, 2).value == 0.1
    # 每条管线一行
    assert ws.cell(7, 1).value == "P1"
    assert ws.cell(7, 2).value == 0.1
    assert ws.cell(8, 1).value == "P2"
    assert ws.cell(8, 2).value == 0.2
